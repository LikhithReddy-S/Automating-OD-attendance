import tkinter as tk
from tkinter import messagebox
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load Excel data
try:
    df = pd.read_excel("student_list.xlsx")
    df.columns = df.columns.str.strip()
    df['Roll Number'] = df['Roll Number'].astype(str).str.strip()
    df['Name'] = df['Name'].astype(str).str.strip()
except Exception as e:
    messagebox.showerror("Error", f"Error loading student_list.xlsx: {e}")
    df = pd.DataFrame(columns=["Roll Number", "Name"])

entries = []

def get_name_by_roll(roll):
    roll = roll.strip()
    result = df[df["Roll Number"].str.strip().str.upper() == roll.upper()]
    if not result.empty:
        return result.iloc[0]["Name"]
    return ""

def add_entry():
    roll = roll_entry.get().strip()
    name = name_entry.get().strip()
    slot_input = slot_entry.get().strip()
    event = event_entry.get().strip()

    if not roll or not name or not slot_input or not event:
        messagebox.showerror("Error", "All fields are required.")
        return

    try:
        slots = [int(s.strip()) for s in slot_input.split(",") if s.strip().isdigit()]
    except:
        messagebox.showerror("Error", "Invalid slot format. Use comma-separated numbers like 1,2,3.")
        return

    for slot in slots:
        entries.append({
            "Roll Number": roll,
            "Name": name,
            "Slot": slot,
            "Event": event,
            "Date": datetime.now().strftime("%d-%m-%Y")
        })

    roll_entry.delete(0, tk.END)
    name_entry.delete(0, tk.END)
    slot_entry.delete(0, tk.END)
    messagebox.showinfo("Success", "Entry added successfully!")

def get_name():
    roll = roll_entry.get().strip()
    name = get_name_by_roll(roll)
    if name:
        name_entry.delete(0, tk.END)
        name_entry.insert(0, name)
    else:
        messagebox.showwarning("Not Found", "Roll Number not found. Enter name manually.")

def export_to_excel():
    if not entries:
        messagebox.showerror("Error", "No entries to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "OD List"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"{event_entry.get().strip()} - OD List"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["S.no", "Roll Number", "Name", "Slot", "Event", "Date"]
    ws.append(headers)

    for i, entry in enumerate(entries, start=1):
        ws.append([
            i,
            entry["Roll Number"],
            entry["Name"],
            entry["Slot"],
            entry["Event"],
            entry["Date"]
        ])

    filename = f"{event_entry.get().strip().replace(' ', '_')}_OD_List_{datetime.now().strftime('%d%m%Y')}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Success", f"Excel saved as {filename}")

# GUI
root = tk.Tk()
root.title("OD Entry App")
root.geometry("500x400")

tk.Label(root, text="Roll Number:").pack()
roll_entry = tk.Entry(root)
roll_entry.pack()

tk.Button(root, text="Get Name", command=get_name).pack(pady=2)

tk.Label(root, text="Name:").pack()
name_entry = tk.Entry(root)
name_entry.pack()

tk.Label(root, text="Event Name:").pack()
event_entry = tk.Entry(root)
event_entry.pack()

tk.Label(root, text="Enter Slots (e.g. 2,3,4):").pack()
slot_entry = tk.Entry(root)
slot_entry.pack()

tk.Button(root, text="Add Entry", command=add_entry).pack(pady=10)
tk.Button(root, text="Export to Excel", command=export_to_excel).pack(pady=10)

root.mainloop()
